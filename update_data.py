#!/usr/bin/env python3
"""
Parse new Excel and bake data into index.html + update data.json
"""
import json, re, sys
import openpyxl
from datetime import datetime

if len(sys.argv) < 2:
    print("Usage: python3 update_data.py <excel_path>"); sys.exit(1)
XL_PATH   = sys.argv[1]
HTML_PATH = sys.path[0] + '/index.html' if sys.path[0] else 'index.html'
JSON_PATH = sys.path[0] + '/data.json'  if sys.path[0] else 'data.json'
# Allow overrides via argv
if len(sys.argv) >= 3: HTML_PATH = sys.argv[2]
if len(sys.argv) >= 4: JSON_PATH = sys.argv[3]

# ─────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────
def find_sh(wb, keywords):
    for kw in keywords:
        for name in wb.sheetnames:
            if kw in name:
                return name
    return None

def rows(ws):
    return list(ws.iter_rows(values_only=True))

def sf(v):
    if v is None: return ''
    return str(v).strip()

def pf(v):
    try: return float(str(v).replace(',','')) if v is not None else 0.0
    except: return 0.0

def pi(v):
    try: return int(pf(v))
    except: return 0

# ─────────────────────────────────────────
# Load workbook
# ─────────────────────────────────────────
print(f"Loading {XL_PATH} ...")
wb = openpyxl.load_workbook(XL_PATH, data_only=True)
print("Sheets:", wb.sheetnames)

s_inv   = find_sh(wb, ['在庫'])
s_pn    = find_sh(wb, ['品名健康度'])
s_sales = find_sh(wb, ['銷貨數 (總數)', '銷貨數(總數)', '銷貨數'])
s_learn = find_sh(wb, ['學習商品(claude)', '學習商品(Claude)'])
s_r03   = find_sh(wb, ['R03022'])
s_sp    = find_sh(wb, ['售價'])
s_erp   = find_sh(wb, ['ERP學習商品生產數量', 'ERP學習商品'])
s_res   = find_sh(wb, ['FY26②保留商品-部門別使用數量', 'FY26②保留商品'])

print(f"  在庫: {s_inv}  品名健康度: {s_pn}  銷貨數: {s_sales}  學習商品: {s_learn}")
print(f"  R03022: {s_r03}  售價: {s_sp}  ERP: {s_erp}  FY26: {s_res}")

if not s_inv or not s_pn or not s_sales:
    print("ERROR: missing required sheets"); sys.exit(1)

# ─────────────────────────────────────────
# 在庫 sheet → INV + dataDate
# ─────────────────────────────────────────
w_inv = rows(wb[s_inv])
ds = sf(w_inv[2][0]) if len(w_inv) > 2 and w_inv[2] else ''
rc = re.search(r'(\d+)/(\d+)/(\d+)', ds)
if rc:
    y = int(rc.group(1))
    yy = y + 1911 if y < 200 else y
    data_date = f"{yy}/{rc.group(2)}/{rc.group(3)}"
else:
    data_date = datetime.now().strftime('%Y/%m/%d')
print(f"  dataDate: {data_date}")

INV = {}
for r in w_inv[4:]:
    if not r or not r[0]: continue
    c = sf(r[0])
    if not c: continue
    if c not in INV:
        INV[c] = {'qty': 0.0, 'val': 0.0, 'cost': 0.0,
                  'invEv': '', 'gift': False, 'reduce': False,
                  'prodMethod': '', 'dept': ''}
    INV[c]['qty']  += pf(r[3])
    INV[c]['val']  += pf(r[4])
    uc = pf(r[2])
    if uc > 0: INV[c]['cost'] = uc
    if not INV[c]['invEv'] and r[7]: INV[c]['invEv'] = sf(r[7])
    if r[8]: INV[c]['gift'] = True
    if r[9]: INV[c]['reduce'] = True
    pm = sf(r[11]) if len(r) > 11 else ''
    if not INV[c]['prodMethod'] and pm and pm != '-': INV[c]['prodMethod'] = pm
    dp = sf(r[12]) if len(r) > 12 else ''
    if not INV[c]['dept'] and dp: INV[c]['dept'] = dp

print(f"  INV items: {len(INV)}")

# ─────────────────────────────────────────
# 銷貨數(總數) sheet → SALES + MONTHS
# ─────────────────────────────────────────
w_s = rows(wb[s_sales])
h_row = list(w_s[2]) if len(w_s) > 2 else []
MONTHS = []
m_idx = {}
for i, v in enumerate(h_row):
    sv = sf(v)
    if re.match(r'^20\d{4}$', sv):
        MONTHS.append(sv)
        m_idx[sv] = i
LATEST_MONTH = MONTHS[-1] if MONTHS else ''
print(f"  MONTHS: {MONTHS[0]}..{LATEST_MONTH} ({len(MONTHS)} months)")

SALES = {}
for r in w_s[3:]:
    if not r or not r[1]: continue
    c = sf(r[1])
    if not c: continue
    ev_s = sf(r[0])
    nm_s = sf(r[2])
    if c not in SALES:
        SALES[c] = {'ev': ev_s, 'name': nm_s, 'm': {}}
    for m in MONTHS:
        idx = m_idx[m]
        v = pf(r[idx]) if idx < len(r) else 0.0
        SALES[c]['m'][m] = SALES[c]['m'].get(m, 0.0) + v

print(f"  SALES items: {len(SALES)}")

# EV_FROM_SALES
EV_FROM_SALES = {c: SALES[c]['ev'] for c in SALES if SALES[c].get('ev')}

# ─────────────────────────────────────────
# R03022 → GIFT_CODES (set)
# ─────────────────────────────────────────
GIFT_CODES = set()
PRICES = {}

if s_r03:
    w_r = rows(wb[s_r03])
    for r in w_r[6:]:
        if not r or not r[1]: continue
        c = sf(r[1])
        p = pf(r[5]) if len(r) > 5 else 0
        kind = sf(r[4]) if len(r) > 4 else ''
        if kind == '禮物' or p == 99999:
            GIFT_CODES.add(c)

# ─────────────────────────────────────────
# 售價 sheet → PRICES
# ─────────────────────────────────────────
if s_sp:
    w_sp = rows(wb[s_sp])
    for r in w_sp[1:]:
        if not r or not r[0]: continue
        c = sf(r[0])
        p = pf(r[4]) if len(r) > 4 else 0
        kind = sf(r[3]) if len(r) > 3 else ''
        if kind == '禮物' or p == 99999:
            GIFT_CODES.add(c)
            continue
        if p > 0: PRICES[c] = p

print(f"  PRICES: {len(PRICES)}  GIFT_CODES: {len(GIFT_CODES)}")

# ─────────────────────────────────────────
# ERP學習商品生産數量 → PRODUCTION
# ─────────────────────────────────────────
PRODUCTION = {}
if s_erp:
    w_erp = rows(wb[s_erp])
    erp_hdr_idx = -1
    col_code = col_year = col_date = col_qty = -1
    for hi in range(min(10, len(w_erp))):
        hr = list(w_erp[hi] or [])
        try:
            ci = hr.index('產品品號')
            erp_hdr_idx = hi
            col_code = ci
            col_year = hr.index('年度') if '年度' in hr else -1
            col_date = next((i for i, v in enumerate(hr)
                             if v and str(v).replace(' ','') == '實完工'), -1)
            col_qty  = next((i for i, v in enumerate(hr)
                             if v and '已生產量' in str(v).replace(' ','')), -1)
            break
        except ValueError:
            continue
    if erp_hdr_idx >= 0:
        for r in w_erp[erp_hdr_idx+1:]:
            if not r or not r[col_code]: continue
            c   = sf(r[col_code])
            yr  = sf(r[col_year]) if col_year >= 0 and col_year < len(r) else ''
            roc = sf(r[col_date]) if col_date >= 0 and col_date < len(r) else ''
            if isinstance(r[col_date] if col_date >= 0 else None, datetime):
                dt_obj = r[col_date]
                ad_date = dt_obj.strftime('%Y/%m/%d')
            else:
                ad_date = re.sub(r'^(\d+)/', lambda m: str(int(m.group(1))+1911)+'/', roc)
            qty_raw = sf(r[col_qty]) if col_qty >= 0 and col_qty < len(r) else '0'
            qty = pf(qty_raw)
            if not c or not qty: continue
            if c not in PRODUCTION: PRODUCTION[c] = []
            PRODUCTION[c].append({'year': yr, 'date': ad_date, 'qty': qty})
    print(f"  PRODUCTION codes: {len(PRODUCTION)}")

# ─────────────────────────────────────────
# 學習商品(claude) → LEARN + LEARN_CODES
# ─────────────────────────────────────────
LEARN = {}
LEARN_CODES = []
if s_learn:
    w_l = rows(wb[s_learn])
    if not w_l: pass
    else:
        headers = [sf(v) for v in w_l[0]]
        for r in w_l[1:]:
            if not r: continue
            rn = {headers[i].strip(): sf(r[i]) for i in range(min(len(headers), len(r)))}
            id_ = rn.get('商品編號', '').strip()
            if not id_: continue
            LEARN[id_] = {
                'lead':      pi(rn.get('追加製作期_月', 6)) or 6,
                'first':     rn.get('首月銷售月份', '').strip(),
                'price':     pf(rn.get('單價', 0)),
                'prodCount': pi(rn.get('製作数', rn.get('製作數', 0))),
                'inDate':    rn.get('入庫日', rn.get('入庫月', '')).strip(),
                'note':      rn.get('生產動態', '').strip()
            }
            LEARN_CODES.append(id_)
    print(f"  LEARN: {len(LEARN)} codes: {LEARN_CODES}")

# ─────────────────────────────────────────
# 品名健康度 → ALL_ITEMS
# ─────────────────────────────────────────
w_pn = rows(wb[s_pn])
ALL_ITEMS = []
pn_codes = set()
for r in w_pn[4:]:
    if not r or not r[2] or not r[11]: continue
    code = sf(r[2])
    pn_codes.add(code)
    inv0 = INV.get(code, {})
    ALL_ITEMS.append({
        'ev':   sf(r[0]), 'ev2': sf(r[1]),
        'code': code,     'name': sf(r[3]),
        'qty':  pi(r[4]), 'val': pi(r[5]),
        's3':   pf(r[8]), 'avg': pf(r[9]),
        'health': sf(r[11]),
        'invEv':     inv0.get('invEv',''),
        'gift':      inv0.get('gift', False),
        'reduce':    inv0.get('reduce', False),
        'prodMethod':inv0.get('prodMethod',''),
        'dept':      inv0.get('dept','')
    })

# Supplement items not in 品名健康度 (Mirafeel, etc.)
learn_set = set(LEARN_CODES)
for code, inv1 in INV.items():
    if code in pn_codes: continue
    if inv1['qty'] <= 0: continue
    ev = EV_FROM_SALES.get(code, '')
    if not ev: continue
    if code in learn_set: continue
    name = SALES[code]['name'] if code in SALES else code
    ALL_ITEMS.append({
        'ev': ev, 'ev2': ev,
        'code': code, 'name': name,
        'qty':  round(inv1['qty']), 'val': round(inv1['val']),
        's3': 0, 'avg': 0,
        'health': '未評估',
        'invEv':     inv1.get('invEv',''),
        'gift':      inv1.get('gift', False),
        'reduce':    inv1.get('reduce', False),
        'prodMethod':inv1.get('prodMethod',''),
        'dept':      inv1.get('dept','')
    })

print(f"  ALL_ITEMS: {len(ALL_ITEMS)}")

# ─────────────────────────────────────────
# FY26②保留商品 → RESERVE_DATA
# ─────────────────────────────────────────
RESERVE_DATA = {}
if s_res:
    w_res = rows(wb[s_res])
    r2res = list(w_res[1]) if len(w_res) > 1 else []
    r4res = list(w_res[3]) if len(w_res) > 3 else []
    res_cols = []
    for j in range(3, len(r2res)):
        dc = r2res[j]
        dt = r4res[j] if j < len(r4res) else None
        if dc and isinstance(dc, str) and dc.startswith('H') and isinstance(dt, datetime):
            mm = str(dt.month).zfill(2)
            res_cols.append({'j': j, 'dept': dc, 'mo': f"{dt.year}{mm}"})
    for rr in w_res[4:]:
        if not rr or not rr[0]: continue
        rcode = sf(rr[0])
        entry = {}
        for col in res_cols:
            v = rr[col['j']] if col['j'] < len(rr) else None
            try: fv = int(float(str(v))) if v is not None else 0
            except: fv = 0
            if fv > 0:
                dept = col['dept']
                mo   = col['mo']
                if dept not in entry: entry[dept] = {}
                entry[dept][mo] = entry[dept].get(mo, 0) + fv
        if entry: RESERVE_DATA[rcode] = entry
    print(f"  RESERVE_DATA codes: {len(RESERVE_DATA)}")

# ─────────────────────────────────────────
# Keep DIMS and SALE_GROUPS from existing data.json (not from Excel)
# ─────────────────────────────────────────
with open(JSON_PATH) as f:
    old = json.load(f)
DIMS        = old.get('dims', {})
SALE_GROUPS = old.get('saleGroups', {})

# ─────────────────────────────────────────
# Write data.json
# ─────────────────────────────────────────
new_data = {
    'dataDate':    data_date,
    'months':      MONTHS,
    'latestMonth': LATEST_MONTH,
    'inv':         INV,
    'sales':       SALES,
    'learn':       LEARN,
    'learnCodes':  LEARN_CODES,
    'allItems':    ALL_ITEMS,
    'prices':      PRICES,
    'evFromSales': EV_FROM_SALES,
    'giftCodes':   list(GIFT_CODES),
    'production':  PRODUCTION,
    'dims':        DIMS,
    'saleGroups':  SALE_GROUPS
}
with open(JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(new_data, f, ensure_ascii=False, separators=(',',':'))
print(f"  ✓ data.json written")

# ─────────────────────────────────────────
# Update index.html — replace BAKE lines
# ─────────────────────────────────────────
with open(HTML_PATH, 'r', encoding='utf-8') as f:
    lines = f.readlines()

def js(v):
    return json.dumps(v, ensure_ascii=False, separators=(',',':'))

def line_replace(lines, prefix, new_content):
    """Replace the entire line that starts with `prefix`"""
    for i, ln in enumerate(lines):
        if ln.startswith(prefix):
            lines[i] = new_content + '\n'
            return True
    return False

# MONTHS
line_replace(lines, '/*BAKE_MONTHS*/',
    f"/*BAKE_MONTHS*/var MONTHS={js(MONTHS)}, LATEST_MONTH='{LATEST_MONTH}';")

# INV
line_replace(lines, '/*BAKE_INV*/',
    f"/*BAKE_INV*/var INV={js(INV)};")

# SALES
line_replace(lines, '/*BAKE_SALES*/',
    f"/*BAKE_SALES*/var SALES={js(SALES)};")

# LEARN
line_replace(lines, '/*BAKE_LEARN*/',
    f"/*BAKE_LEARN*/var LEARN={js(LEARN)};")

# LEARN_CODES
line_replace(lines, '/*BAKE_LC*/',
    f"/*BAKE_LC*/var LEARN_CODES={js(LEARN_CODES)};")

# ALL_ITEMS
line_replace(lines, '/*BAKE_AI*/',
    f"/*BAKE_AI*/var ALL_ITEMS={js(ALL_ITEMS)};")

# EV_FROM_SALES
line_replace(lines, '/*BAKE_EV*/',
    f"/*BAKE_EV*/var EV_FROM_SALES={js(EV_FROM_SALES)};")

# PRICES (no BAKE marker, starts with 'var PRICES=')
line_replace(lines, 'var PRICES=',
    f"var PRICES={js(PRICES)};")

# GIFT_CODES (no BAKE marker)
gc_list = sorted(GIFT_CODES)
line_replace(lines, 'var GIFT_CODES=',
    f"var GIFT_CODES=new Set({js(gc_list)});")

# RESERVE_DATA
line_replace(lines, '/*BAKE_RESERVE*/',
    f"/*BAKE_RESERVE*/var RESERVE_DATA={js(RESERVE_DATA)};")

# PRODUCTION (has /*BAKE_PRODUCTION_START*/ ... /*BAKE_PRODUCTION_END*/ on same line)
line_replace(lines, '/*BAKE_PRODUCTION_START*/',
    f"/*BAKE_PRODUCTION_START*/var PRODUCTION={js(PRODUCTION)};/*BAKE_PRODUCTION_END*/")

# upStatus span
n_items = len(ALL_ITEMS)
months_cnt = len(MONTHS)
for i, ln in enumerate(lines):
    if 'id="upStatus"' in ln:
        lines[i] = re.sub(
            r'(<span[^>]*id="upStatus"[^>]*>)[^<]*(</span>)',
            f'\\1載入成功 {n_items} 筆｜{data_date}｜{months_cnt}個月資料\\2',
            ln
        )
        break

# dataDate span
for i, ln in enumerate(lines):
    if 'id="dataDate"' in ln:
        lines[i] = re.sub(
            r'(<span[^>]*id="dataDate"[^>]*>)[^<]*(</span>)',
            f'\\1在庫日期：{data_date}\\2',
            ln
        )
        break

with open(HTML_PATH, 'w', encoding='utf-8') as f:
    f.writelines(lines)
print(f"  ✓ index.html updated")
print(f"\nDone! dataDate={data_date}, items={n_items}, months={len(MONTHS)}")
